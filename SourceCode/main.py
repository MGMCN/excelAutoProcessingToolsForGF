import openpyxl
import os

duplicateMode = False  # 做单个sample设置为True

targetToSampleAndCq = {}

targetSet = set()


def readSheet(book, sheet_name):
    sheet = book[sheet_name]

    # 获取 Target、 Sample、 Mean Cq的列坐标
    columns = list(sheet.columns)
    originalTargets = ('Target', 'Sample', 'Cq',)  # 'Cq Mean' ??? 还有啥取名儿啊 ???
    targetsToColumnIndex = {}
    for column in columns:
        for first_cell in column:
            coordinate = first_cell.coordinate
            value = sheet[coordinate].value
            if value in originalTargets:
                targetsToColumnIndex[value] = coordinate.__getitem__(0)
                break

    # 生成对应Target Sample Mean列的列数据并且对应存起来
    targetsToColumn = {}
    for target in targetsToColumnIndex:
        targetsToColumn[target] = list(sheet[targetsToColumnIndex[target]])

    # 生成 Target->Sample 的字典映射 targetToSampleAndCq['Target'] = {} Target有Actin和例如SOMT9、IOMT4、OMT38之类的
    targetToSampleAndCq = {}
    cnt = 0
    global targetSet
    for Target, Sample, Cq in zip(targetsToColumn['Target'], targetsToColumn['Sample'], targetsToColumn['Cq']):
        tv = Target.value
        sv = Sample.value
        cqv = Cq.value
        if tv is None:
            continue
        if cnt == 0:  # 仅仅是为了跳过第一行而设置的flag
            cnt += 1
            continue
        targetSet.add(tv)
        if tv not in targetToSampleAndCq:
            targetToSampleAndCq[tv] = {sv: cqv}
        else:
            targetToSampleAndCq[tv][sv] = cqv

    return targetToSampleAndCq


def copyFromReturnValue(tsc):
    global targetToSampleAndCq
    for tv in tsc:
        for sv in tsc[tv]:
            if tv not in targetToSampleAndCq:
                targetToSampleAndCq[tv] = {sv: tsc[tv][sv]}
            else:
                targetToSampleAndCq[tv][sv] = tsc[tv][sv]


def readExcelFiles(dir_path):
    fileList = os.listdir(dir_path)
    for file in fileList:
        if '.xlsx' not in file or 'output' in file:
            continue
        print("当前我们正在访问 ->", file)
        book = openpyxl.load_workbook(dir_path + '/' + file)  # 这个/对应MacOS的格式
        sheet_names = book.sheetnames
        if len(sheet_names) == 1:
            tsc = readSheet(book, sheet_name=sheet_names.__getitem__(0))
            copyFromReturnValue(tsc)
            continue
        print("当前文件有sheet ->", sheet_names)
        print("当你需要读取某一个sheet的时候，请直接输入那个sheet名(全部都读取请输入all),跳过访问当前文件请输入nothing")
        sheet_name = input()
        if sheet_name == 'all':
            # for loop 处理所有sheet
            for sn in sheet_names:
                tsc = readSheet(book, sheet_name=sn)
                copyFromReturnValue(tsc)
        elif sheet_name == 'nothing':
            continue
        else:
            # 处理单个sheet
            tsc = readSheet(book, sheet_name=sheet_name)
            copyFromReturnValue(tsc)


# 每一组的数据处理函数
def calculate(actins, samples, target):
    ans = {}
    sampleCqMean = 0.0
    actinCqMean = 0.0
    size = len(samples)
    for s in samples:
        ans[s['Sample']] = {}  # 初始化3个样本 或者 1个样本
        ans[s['Sample']][target] = {}
        ans[s['Sample']]['Actin'] = {}
        sampleCqMean += s['Cq']
        actinCqMean += actins[s['Sample']]
    for s in samples:
        ans[s['Sample']][target]['CqMean'] = sampleCqMean / size
        ans[s['Sample']]['Actin']['CqMean'] = actinCqMean / size
    for s in samples:
        ans[s['Sample']]['△Cq'] = ans[s['Sample']]['Actin']['CqMean'] - ans[s['Sample']][target]['CqMean']
    for s in samples:
        ans[s['Sample']]['2△Cq'] = 2 ** ans[s['Sample']]['△Cq']
    _2deltaCqMean = 0
    for s in samples:
        _2deltaCqMean += ans[s['Sample']]['2△Cq']
    for s in samples:
        ans[s['Sample']]['2△CqMean'] = _2deltaCqMean / size
    return ans


def chooseSample(choosed):
    global targetToSampleAndCq
    choosedList = []
    for Sample in targetToSampleAndCq[choosed]:
        # print(Sample)
        choosedList.append({'Sample': Sample, 'Cq': targetToSampleAndCq[choosed][Sample]})
    choosedList = sorted(choosedList, key=lambda x: x['Sample'])

    # 当duplicateMode为True的时候pace=1，此处因为给的xlsx为单个样本所以设置为pace=1
    pace = 1
    cnt = 0
    calculateDataSet = []
    ans = None
    ansList = []
    print('choosed ->', choosed)
    for Sample in choosedList:
        # print(Sample)
        # print(calculateDataSet, len(calculateDataSet))
        if cnt < pace:
            calculateDataSet.append(Sample)
            cnt += 1
        else:
            # 计算一组数据
            ans = calculate(actins=targetToSampleAndCq['Actin'], samples=calculateDataSet, target=choosed)
            ansList.append(ans)
            # print(ans)
            # 计算完了后添加这次遍历到的Sample
            calculateDataSet = []
            cnt = 1
            calculateDataSet.append(Sample)

    # 这儿还得处理最后一次，因为最后一次for loop就终止了
    ans = calculate(actins=targetToSampleAndCq['Actin'], samples=calculateDataSet, target=choosed)
    ansList.append(ans)
    # print(ans)
    return ansList


def output(target, samples):
    if not os.path.exists('output.xlsx'):
        f = openpyxl.Workbook()
    else:
        f = openpyxl.load_workbook('output.xlsx')

    f.create_sheet(target)
    sheet = f[target]

    sheet.cell(row=1, column=1, value='Target')
    sheet.cell(row=1, column=2, value='Sample')
    sheet.cell(row=1, column=3, value='Cq')
    sheet.cell(row=1, column=4, value='2△CqMean')

    row = 2
    col = 1

    for sample in samples:
        for sampleName in sample:
            sheet.cell(row=row, column=col, value='Actin')
            col += 1
            sheet.cell(row=row, column=col, value=sampleName)
            col += 1
            sheet.cell(row=row, column=col, value=sample[sampleName]['Actin']['CqMean'])
            col += 1
            sheet.cell(row=row, column=col, value=sample[sampleName]['2△CqMean'])

            row += 1
            col = 1
            sheet.cell(row=row, column=col, value=target)
            col += 1
            sheet.cell(row=row, column=col, value=sampleName)
            col += 1
            sheet.cell(row=row, column=col, value=sample[sampleName][target]['CqMean'])
        row += 1
        col = 1

    f.save("output.xlsx")
    f.close()


if __name__ == "__main__":
    readExcelFiles('.')
    print(targetSet)
    for target in targetSet:
        if target == 'Actin':
            continue
        ansList = chooseSample(target)
        output(target, ansList)
